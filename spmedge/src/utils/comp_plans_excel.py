import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os


def load_json_files(file_paths):
    """Load JSON files and return as list of dictionaries."""
    data = []
    for file_path in file_paths:
        with open(file_path, 'r') as f:
            data.append(json.load(f))
    return data


def create_plans_summary(data):
    """Create plans summary DataFrame."""
    rows = []
    for plan in data:
        plan_info = plan["plan_info"]
        row = {
            "Role": plan_info.get("role", ""),
            "Business Unit": plan_info.get("business_unit", ""),
            "Plan ID": plan_info.get("plan_id", ""),
            "Plan Year": plan_info.get("plan_year", ""),
            "Start Date": plan.get("effective_dates", {}).get("start_date", ""),
            "End Date": plan.get("effective_dates", {}).get("end_date", ""),
            "Region": plan_info.get("region", "")
        }
        rows.append(row)
    
    return pd.DataFrame(rows)


def create_compensation_components(data):
    """Create compensation components DataFrame."""
    rows = []
    for plan in data:
        role = plan["plan_info"].get("role", "")
        for comp in plan.get("compensation_components", []):
            row = {
                "Role": role,
                "Component Name": comp.get("name", ""),
                "Type": comp.get("type", ""),
                "Category": comp.get("category", ""),
                "Frequency": comp.get("frequency", ""),
                "Target Amount": comp.get("target_amount", ""),
                "Structure": comp.get("structure", "")
            }
            rows.append(row)
    
    return pd.DataFrame(rows)


def create_special_provisions(data):
    """Create special provisions DataFrame."""
    rows = []
    for plan in data:
        role = plan["plan_info"].get("role", "")
        for provision in plan.get("special_provisions", []):
            row = {
                "Role": role,
                "Provision Name": provision.get("name", ""),
                "Description": provision.get("description", ""),
                "Conditions": provision.get("conditions", ""),
                "SPM Component": provision.get("spm_mapping", {}).get("spm_component", "")
            }
            rows.append(row)
    
    return pd.DataFrame(rows)


def create_terms_and_conditions(data):
    """Create terms and conditions data structure."""
    terms_data = {}
    for plan in data:
        role = plan["plan_info"].get("role", "")
        terms_data[role] = []
        
        for term in plan.get("terms_and_conditions", []):
            component_type = term.get("component_type", "Terms & Conditions")
            description = term.get("description", "")
            terms_data[role].append({
                "type": component_type,
                "description": description
            })
    
    return terms_data


def apply_style_to_worksheet(ws, has_header=True):
    """Apply styling to worksheet."""
    # Define styles
    header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Apply styles to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            if has_header and cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        ws.column_dimensions[column].width = adjusted_width


def create_terms_sheet(wb, terms_data):
    """Create a formatted terms and conditions sheet."""
    ws = wb.create_sheet(title="Terms & Conditions")
    
    # Add title
    ws.append(["Terms & Conditions"])
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add content for each role
    row_idx = 3
    for role, terms in terms_data.items():
        # Add role name
        ws.append([role])
        ws.merge_cells(f'A{row_idx}:E{row_idx}')
        role_cell = ws[f'A{row_idx}']
        role_cell.font = Font(bold=True, size=12)
        
        row_idx += 1
        
        # Add terms for this role
        for term in terms:
            term_type = term.get("type", "")
            description = term.get("description", "")
            
            ws.append([f"{term_type}:", description])
            ws.merge_cells(f'A{row_idx}:B{row_idx}')
            ws.merge_cells(f'C{row_idx}:E{row_idx}')
            
            # Apply styles
            type_cell = ws[f'A{row_idx}']
            type_cell.font = Font(bold=True)
            
            row_idx += 1
        
        # Add spacing between roles
        ws.append([""])
        row_idx += 1
    
    # Apply borders and other styling
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20


def create_excel_from_json(json_files, output_file="compensation_plans.xlsx"):
    """Create Excel workbook from JSON data."""
    # Load JSON data
    data = load_json_files(json_files)
    
    # Create DataFrames
    plans_df = create_plans_summary(data)
    comp_df = create_compensation_components(data)
    provisions_df = create_special_provisions(data)
    terms_data = create_terms_and_conditions(data)
    
    # Create workbook
    wb = Workbook()
    
    # Create Plans Summary sheet
    ws_plans = wb.active
    ws_plans.title = "Plans Summary"
    for r_idx, row in enumerate(dataframe_to_rows(plans_df, index=False, header=True), 1):
        ws_plans.append(row)
    apply_style_to_worksheet(ws_plans)
    
    # Create Compensation Components sheet
    ws_comp = wb.create_sheet(title="Compensation Components")
    for r_idx, row in enumerate(dataframe_to_rows(comp_df, index=False, header=True), 1):
        ws_comp.append(row)
    apply_style_to_worksheet(ws_comp)
    
    # Create Special Provisions sheet
    ws_prov = wb.create_sheet(title="Special Provisions")
    for r_idx, row in enumerate(dataframe_to_rows(provisions_df, index=False, header=True), 1):
        ws_prov.append(row)
    apply_style_to_worksheet(ws_prov)
    
    # Create Terms and Conditions sheet
    create_terms_sheet(wb, terms_data)
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")


# Example usage
if __name__ == "__main__":
    import sys
    import glob
    
    # Default to data/stage_process directory
    data_dir = "data/stage_process"
    json_pattern = os.path.join(data_dir, "*.json")
    json_files = glob.glob(json_pattern)
    
    if json_files:
        create_excel_from_json(json_files)
    else:
        print(f"Error: No JSON files found in {data_dir}")